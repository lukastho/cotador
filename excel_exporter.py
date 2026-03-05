import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime

def export_to_excel(df, filename_prefix="Relatorio_Precos"):
    """
    Exports the DataFrame to Excel with formatting.
    Items that are opportunities are highlighted in green.
    """
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{filename_prefix}_{date_str}.xlsx"

    # Write to Excel
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Resultados')

    workbook = writer.book
    worksheet = writer.sheets['Resultados']

    # Highlight opportunities in green
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Find the "Oportunidade" column index (0-indexed in DF, 1-indexed in Excel)
    opportunity_col_idx = df.columns.get_loc("Oportunidade") + 1

    for row_idx, is_opportunity in enumerate(df["Oportunidade"], start=2): # Headers are row 1
        if is_opportunity:
            # Highlight the whole row or just some cells? Let's do the whole row.
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = green_fill

    writer.close()
    return filename

if __name__ == "__main__":
    test_df = pd.DataFrame([
        {"Nome do Produto": "Item A", "Loja": "L", "Preço Atual": 100, "Custo Médio (PDF)": 150, "Oportunidade": True},
        {"Nome do Produto": "Item B", "Loja": "L", "Preço Atual": 100, "Custo Médio (PDF)": 90, "Oportunidade": False}
    ])
    f = export_to_excel(test_df, "Test_Report")
    print(f"File created: {f}")
