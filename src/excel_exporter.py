import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime

def export_to_excel(df, filename_prefix="Cotacao_Cacamba"):
    """
    Exporta o DataFrame para um arquivo Excel (.xlsx) com formatação condicional.
    """
    # Geração automática do nome do arquivo com a data atual
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"{filename_prefix}_{date_str}.xlsx"

    # Escrita do arquivo usando o motor openpyxl
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Resultados')

    workbook = writer.book
    worksheet = writer.sheets['Resultados']

    # DEFINIÇÃO DO DESTAQUE VISUAL:
    # Cor verde claro (C6EFCE) para itens marcados como OPORTUNIDADE
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Itera sobre as linhas para aplicar a regra de formatação
    # O Pandas começa na linha 2 no Excel (headers na 1)
    for row_idx, status in enumerate(df["Status_Compra"], start=2):
        if status == 'OPORTUNIDADE DE COMPRA':
            # Aplica o destaque em todas as colunas daquela linha
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = green_fill

    writer.close()
    return filename
